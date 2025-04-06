import difflib
import openpyxl
from openpyxl.styles import PatternFill, Alignment

"""
Reads a file and returns its contents as a list for further processing and comparison.
Args: string consisting of file path.
"""
def read_file(filename):
    with open(filename, 'r') as file:
        return file.read().splitlines()

"""
Compares 2 source code extracts line-by-line to identify differences.
Args:
new_file (string): The file path to the newer version of the source code.
old_file (string): The file path to the older version of the source code.
"""
def compare_files(new_file, old_file):
    diff = list(difflib.ndiff(read_file(old_file), read_file(new_file)))
    old_lines, new_lines = [], []
    old_line_number, new_line_number = 1, 1

    for line in diff:
        if line.startswith(' '):  # Unchanged line
            formatted_line_new = f"{new_line_number} {line[2:].rstrip()}"
            formatted_line_old = f"{old_line_number} {line[2:].rstrip()}"
            new_lines.append(formatted_line_new)
            old_lines.append(formatted_line_old)
            new_line_number += 1
            old_line_number += 1
        elif line.startswith('-'):
            old_lines.append(f"{old_line_number} {line[2:].rstrip()}")
            new_lines.append('')
            old_line_number += 1
        elif line.startswith('+'):
            formatted_line_new = f"{new_line_number} {line[2:].rstrip()}"
            new_lines.append(formatted_line_new)
            old_lines.append('')
            new_line_number += 1

    return new_lines, old_lines

"""
Generates an Excel spreadsheet (.xlsx), with each row in the spreadsheet corresponding to a numbered line from the new/old source code inputs.
Breaks are utilized to maintain alignment between the two scripts for added or removed code.
Color coding (native Excel highlight feature) is utilized to identify changes between the two scripts:
Green (present in the new script) represents LOCs which were removed or modified, and red (present in the old script) represents LOCs which were removed.
Args:
new_lines (list): List of lines from the newer version of the source code.
old_lines (list) List of lines from the older version of the source code.
output_file (string): The file path where the generated Excel file will be saved.
"""
def create_excel(new_lines, old_lines, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'File Comparison'

    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    center_aligned_text = Alignment(horizontal='center')

    for row, (line_new, line_old) in enumerate(zip(new_lines, old_lines), start=1):
        line_new_number, line_new_content = line_new.split(' ', 1) if line_new else ('', '')
        line_old_number, line_old_content = line_old.split(' ', 1) if line_old else ('', '')
        cell_new_number = ws.cell(row, 1, line_new_number)
        cell_new_number.alignment = center_aligned_text
        ws.cell(row, 2, line_new_content)
        cell_old_number = ws.cell(row, 3, line_old_number)
        cell_old_number.alignment = center_aligned_text
        ws.cell(row, 4, line_old_content)
        if line_new_content and not line_old_content:
            ws.cell(row, 2).fill = green_fill
        elif line_old_content and not line_new_content:
            ws.cell(row, 4).fill = red_fill
    column_widths = [0, 0, 0, 0]

    for col in ws.columns:
        for cell in col:
            try:
                column_widths[cell.column - 1] = max(column_widths[cell.column - 1], len(cell.value))
            except TypeError:
                continue

    for i, width in enumerate(column_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = width * 1.2
    wb.save(output_file)

"""
Main execution function for the tool.
Prompts the user for input of the (old and new) source code export files to be compared, and to designate the name of the output file.
Further calls the functions defined above to read, compare, and generate an Excel spreadsheet that highlights the differences between these files.
"""
def main():
    new_file = input("Enter the path of the new file: ").strip('"')
    old_file = input("Enter the path of the old file: ").strip('"')
    output_file = input("Enter the output Excel file name: ")
    new_lines, old_lines = compare_files(new_file, old_file)
    create_excel(new_lines, old_lines, output_file)
    print(f"Comparison Excel file '{output_file}' created successfully.")


if __name__ == "__main__":
    main()
